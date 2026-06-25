"""Excel报告生成"""

from __future__ import annotations

import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from config.settings import RISK_TIERS
from db.models import PipelineResult


TIER_FILLS = {
    "reach": PatternFill(start_color="FF4B4B", end_color="FF4B4B", fill_type="solid"),
    "slight_reach": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
    "match": PatternFill(start_color="00CC66", end_color="00CC66", fill_type="solid"),
    "safe": PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid"),
    "backup": PatternFill(start_color="999999", end_color="999999", fill_type="solid"),
}

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def generate_excel(result: PipelineResult) -> bytes:
    wb = Workbook()

    _write_profile_sheet(wb.active, result)

    for strategy in result.strategies:
        ws = wb.create_sheet(title=strategy.strategy_label)
        _write_strategy_sheet(ws, strategy, result)

    ws_risk = wb.create_sheet(title="风险汇总")
    _write_risk_sheet(ws_risk, result)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_profile_sheet(ws, result: PipelineResult):
    ws.title = "考生信息"
    student = result.student

    ws.append(["志愿质检员 — 审核报告"])
    ws.merge_cells("A1:F1")
    ws["A1"].font = Font(bold=True, size=16)

    ws.append([])
    fields = [
        ("省份", student.province),
        ("分数", student.score),
        ("排名", student.rank),
        ("选考科目", ", ".join(student.all_subjects())),
        ("意向城市", ", ".join(student.preferred_cities)),
        ("意向专业", ", ".join(student.preferred_majors)),
        ("策略偏好", student.risk_preference),
        ("接受调剂", "是" if student.accept_adjustment else "否"),
        ("推荐志愿总数", len(result.all_recommendations)),
    ]
    for label, value in fields:
        ws.append([label, value])

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.border = THIN_BORDER


def _write_strategy_sheet(ws, strategy, result: PipelineResult):
    headers = ["序号", "分层", "院校", "专业", "城市", "学费",
               "历年最低排名", "今年计划", "风险等级", "推荐理由"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    for i, rec in enumerate(strategy.recommendations, 1):
        tier_info = RISK_TIERS.get(rec.tier, {"label": rec.tier})
        ranks = " / ".join(str(r) for r in rec.historical_min_ranks) if rec.historical_min_ranks else "-"
        row_data = [
            i,
            tier_info["label"],
            rec.university_name,
            rec.major_name,
            rec.city or "-",
            rec.tuition or "-",
            ranks,
            rec.current_year_plan_count or "-",
            rec.risk_level,
            rec.reason[:100],
        ]
        ws.append(row_data)
        row_idx = ws.max_row
        tier_fill = TIER_FILLS.get(rec.tier)
        if tier_fill:
            ws.cell(row=row_idx, column=2).fill = tier_fill
            ws.cell(row=row_idx, column=2).font = Font(color="FFFFFF", bold=True)
        for cell in ws[row_idx]:
            cell.border = THIN_BORDER

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 15


def _write_risk_sheet(ws, result: PipelineResult):
    ws.append(["院校", "风险类型", "风险详情", "严重程度"])
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    for uni_key, risks in result.charter_risks.items():
        for risk in risks:
            ws.append([uni_key, risk.risk_type, risk.detail, risk.severity])
            for cell in ws[ws.max_row]:
                cell.border = THIN_BORDER

    if result.audit_result:
        ws.append([])
        ws.append(["审核结果", result.audit_result.status])
        for item in result.audit_result.high_risks:
            ws.append(["高风险", item])
        for item in result.audit_result.warnings:
            ws.append(["警告", item])
